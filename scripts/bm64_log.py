import asyncio
from bleak import BleakClient, BleakScanner
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook

DEVICE_ADDRESS = "A4:C1:38:A5:20:BB"
MEAS_CHAR_UUID = "00002a35-0000-1000-8000-00805f9b34fb"

EXCEL_FILE = "log.xlsx"

measurement_event = asyncio.Event()


def parse_measurement(data: bytes):
    try:
        flags = data[0]
        index = 1

        units_kpa = bool(flags & 0x01)
        timestamp_present = bool(flags & 0x02)
        pulse_present = bool(flags & 0x04)
        user_present = bool(flags & 0x08)
        status_present = bool(flags & 0x10)

        systolic = int.from_bytes(data[index:index+2], 'little')
        index += 2
        diastolic = int.from_bytes(data[index:index+2], 'little')
        index += 2
        _ = int.from_bytes(data[index:index+2], 'little')
        index += 2

        map_val = diastolic + (systolic - diastolic) / 3

        if units_kpa:
            systolic = round(systolic * 7.50062)
            diastolic = round(diastolic * 7.50062)
            map_val *= 7.50062

        map_val = round(map_val, 2)

        if timestamp_present:
            year = int.from_bytes(data[index:index+2], 'little')
            month = data[index+2]
            day = data[index+3]
            hour = data[index+4]
            minute = data[index+5]
            second = data[index+6]
            timestamp = f"{year:04}-{month:02}-{day:02} {hour:02}:{minute:02}:{second:02}"
            index += 7
        else:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        pulse = None
        if pulse_present:
            pulse = int.from_bytes(data[index:index+2], 'little')
            index += 2

        if user_present:
            index += 1
        if status_present:
            index += 2

        return {
            "timestamp": timestamp,
            "systole": systolic,
            "diastole": diastolic,
            "map": map_val,
            "pulse": pulse
        }
    except Exception as e:
        print(f"Parser error: {e}")
        return None


def write_to_excel(values):
    path = os.path.join(os.path.dirname(__file__), EXCEL_FILE)
    headers = ["Zeitstempel", "Systole", "Diastole", "MAP", "Puls"]

    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Messwerte"
        ws.append(headers)
    else:
        wb = load_workbook(path)
        ws = wb.active

    ws.append([
        values["timestamp"],
        values["systole"],
        values["diastole"],
        values["map"],
        values["pulse"] if values["pulse"] is not None else ""
    ])

    wb.save(path)
    print(f"Saved: {values}")


def handle_notification(sender, data):
    values = parse_measurement(data)
    if values:
        write_to_excel(values)
        measurement_event.set()


async def log_once(client):
    measurement_event.clear()
    await client.start_notify(MEAS_CHAR_UUID, handle_notification)
    try:
        await asyncio.wait_for(measurement_event.wait(), timeout=60)
    except asyncio.TimeoutError:
        print("No measurement received")
    await client.stop_notify(MEAS_CHAR_UUID)


async def main():
    while True:
        device = await BleakScanner.find_device_by_address(DEVICE_ADDRESS, timeout=10.0)
        if not device:
            await asyncio.sleep(5)
            continue
        try:
            async with BleakClient(device) as client:
                if client.is_connected:
                    await log_once(client)
        except Exception as e:
            print(f"Connection error: {e}")
        await asyncio.sleep(5)


if __name__ == "__main__":
    asyncio.run(main())
